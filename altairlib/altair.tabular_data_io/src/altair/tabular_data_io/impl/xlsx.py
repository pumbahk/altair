from six import text_type
import tempfile
import shutil
import openpyxl

PREFERRED_MIME_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

class XlsxTabularDataReader(object):
    preferred_mime_type = PREFERRED_MIME_TYPE

    def __init__(self, exts, names):
        self.exts = exts
        self.names = names

    def open(self, f, **options):
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        return (
            [col.value for col in row]
            for row in ws.rows
            )



class XlsxTabularDataWriter(object):
    preferred_mime_type = PREFERRED_MIME_TYPE

    class Helper(object):
        def __init__(self, wb, ws, f, max_inmemory_buf_size):
            self.wb = wb
            self.ws = ws
            self.f = f
            self.max_inmemory_buf_size = max_inmemory_buf_size or 16777216
            self.r = 0

        def close(self):
            if self.wb is not None:
                if not isinstance(self.f, (str, text_type)):
                    maybe_filelike = True
                if maybe_filelike and not hasattr(self.f, 'tell'):
                    with tempfile.SpooledTemporaryFile(max_size=self.max_inmemory_buf_size) as f:
                        self.wb.save(f)
                        f.seek(0)
                        shutil.copyfileobj(f, self.f)
                else:
                    self.wb.save(self.f)
                self.wb = self.ws = None
                if maybe_filelike and hasattr(self.f, 'close'):
                    self.f.close()

        def __call__(self, cols):
            self.ws.append(cols)
            self.r += 1    


    def __init__(self, exts, names):
        self.exts = exts
        self.names = names

    def open(self, f, encoding=u'UTF-8', sheet_name=u'Sheet1', **options):
        wb = openpyxl.Workbook(encoding=encoding, write_only=True)
        ws = wb.create_sheet(None, sheet_name)
        return self.Helper(wb, ws, f, max_inmemory_buf_size=options.get('max_inmemory_buf_size'))
